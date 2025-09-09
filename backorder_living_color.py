import os
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Alignment, numbers
from openpyxl.utils import get_column_letter
from datetime import datetime, timedelta

# --- CONFIGURABLES ---
RAW_KEY_ORDER = ["ORDER #", "ITEM NO"]
USER_COLS = ["PCX DOCK", "COMMENTS"]
COLOR_KEYWORDS = {
    "PROCESS": "DEEAD0",
    "PICKUP": "C5D9F1",
    "ISSUE": "FFFF00",
    "TESTING": "CCC0DA",
    "SCHEDULED": "F2DCDB",
}

REPORT_HEADERS = [
    "ORDER #", "CUST PO", "ORDER DATE", "PCX DOCK", "ITEM NO", "MFG", "HIP ASA", "UNIT PRICE", "UNIT COST",
    "CUST NAME", "SALESMAN NAME", "DUE DATE", "STOCK", "GP UNIT", "GP TOTAL", "TOTAL SALE", "COMMENTS"
]

def find_previous_report(history_dir="report_history"):
    today = datetime.now()
    for delta in range(1, 8):  # look back up to 7 days
        prev_date = (today - timedelta(days=delta)).strftime("%m%d%y")
        for fname in os.listdir(history_dir):
            if prev_date in fname and fname.lower().endswith(".xlsx"):
                return os.path.join(history_dir, fname)
    return None

def load_prev_userfields(prev_report_file):
    # Load both sheets if present, concat
    if not prev_report_file or not os.path.exists(prev_report_file):
        return pd.DataFrame()
    dfs = []
    for sh in ["MILITARY", "COMMERCIAL"]:
        try:
            df = pd.read_excel(prev_report_file, sheet_name=sh)
            dfs.append(df)
        except Exception:
            continue
    if not dfs:
        return pd.DataFrame()
    prev = pd.concat(dfs, axis=0, ignore_index=True)
    for col in RAW_KEY_ORDER:
        prev[col] = prev[col].astype(str).str.strip()
    # Only keep keys + user cols:
    prev = prev[RAW_KEY_ORDER + USER_COLS].drop_duplicates(RAW_KEY_ORDER, keep="last")
    prev = prev.set_index(RAW_KEY_ORDER)
    return prev

def color_row(ws, row_num, color_hex):
    for col in range(1, ws.max_column+1):
        ws.cell(row=row_num, column=col).fill = PatternFill(start_color=color_hex, end_color=color_hex, fill_type="solid")

def main(
    raw_data_file,
    output_file=None,
    history_dir="report_history"
):
    # --- Read RAW DATA ---
    raw_df = pd.read_excel(raw_data_file)
    # Rename columns as needed to match your living report
    col_map = {
        "order_no": "ORDER #",
        "cust_po": "CUST PO",
        "order_dt": "ORDER DATE",
        "item_no": "ITEM NO",
        "manu_no": "MFG",
        "ship_asap": "HIP ASA",
        "unit_price": "UNIT PRICE",
        "unit_cost": "UNIT COST",
        "cust_name": "CUST NAME",
        "slsman_name": "SALESMAN NAME",
        "due_date": "DUE DATE",
        "from_stk": "STOCK",
    }
    raw_df.columns = [c.strip().upper() for c in raw_df.columns]
    for k, v in col_map.items():
        if k.upper() in raw_df.columns:
            raw_df.rename(columns={k.upper(): v}, inplace=True)
    # Ensure all needed columns are there
    for col in REPORT_HEADERS:
        if col not in raw_df.columns:
            if col in USER_COLS:
                raw_df[col] = ""
            else:
                raw_df[col] = ""
    # Fix key columns as string type and strip
    for col in RAW_KEY_ORDER:
        raw_df[col] = raw_df[col].astype(str).str.strip()
    raw_df = raw_df.drop_duplicates(subset=RAW_KEY_ORDER)

    # --- Load previous report user fields ---
    prev_report = find_previous_report(history_dir)
    prev_users = load_prev_userfields(prev_report) if prev_report else pd.DataFrame()

    # --- Carry over PCX DOCK and COMMENTS ---
    if not prev_users.empty:
        raw_df.set_index(RAW_KEY_ORDER, inplace=True)
        for idx in raw_df.index:
            if idx in prev_users.index:
                for col in USER_COLS:
                    val = prev_users.loc[idx, col]
                    if pd.notna(val):
                        raw_df.at[idx, col] = val
        raw_df.reset_index(inplace=True)

    # --- Order columns for export ---
    raw_df = raw_df[REPORT_HEADERS]

    # --- Split MILITARY/COMMERCIAL ---
    is_military = raw_df["SALESMAN NAME"].str.upper() == "MANUEL ORTEGA"
    military = raw_df[is_military].copy()
    commercial = raw_df[~is_military].copy()

    # --- Write to Excel with color coding ---
    if not output_file:
        output_file = f"BACKORDER REPORT {datetime.now().strftime('%m%d%y')}.xlsx"
    wb = Workbook()
    for sheetname, df in [("MILITARY", military), ("COMMERCIAL", commercial)]:
        ws = wb.create_sheet(sheetname)
        ws.append(REPORT_HEADERS)
        # Write data
        for r, (_, row) in enumerate(df.iterrows(), start=2):
            ws.append(list(row))
            # Color row if COMMENTS contains a keyword
            comment = str(row["COMMENTS"]).upper()
            for k, color in COLOR_KEYWORDS.items():
                if k in comment:
                    color_row(ws, r, color)
                    break
        # Header styling
        for col in range(1, len(REPORT_HEADERS)+1):
            cell = ws.cell(row=1, column=col)
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center")
        # Currency formatting
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=8, max_col=10):
            for cell in row:
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
        # Date formatting (ORDER DATE, DUE DATE)
        for colname in ["ORDER DATE", "DUE DATE"]:
            colnum = REPORT_HEADERS.index(colname) + 1
            for row in range(2, ws.max_row+1):
                cell = ws.cell(row=row, column=colnum)
                if cell.value:
                    try:
                        cell.value = pd.to_datetime(cell.value).strftime("%m-%d-%y")
                    except Exception:
                        pass
                cell.number_format = "MM-DD-YY"
        # Set column widths
        for i, col in enumerate(REPORT_HEADERS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = max(12, len(col)+2)
        ws.freeze_panes = "A2"
    # Remove default sheet
    wb.remove(wb["Sheet"])
    wb.save(output_file)
    print(f"Saved: {output_file}")

if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser(description="Backorder Living Report Maker (with color coding)")
    parser.add_argument("raw_data_file", help="Raw data Excel file (xlsx)")
    parser.add_argument("--output", help="Output file (xlsx)")
    parser.add_argument("--history", default="report_history", help="Report history dir")
    args = parser.parse_args()
    main(args.raw_data_file, args.output, args.history)
