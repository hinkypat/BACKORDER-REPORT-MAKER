"""
Excel Report Generation Module
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment
from openpyxl.chart import BarChart, PieChart, LineChart, Reference
from openpyxl.utils.dataframe import dataframe_to_rows
import logging
from datetime import datetime

class ExcelGenerator:
    def __init__(self, config):
        self.config = config
        self.logger = logging.getLogger(__name__)
        
        # Define styling
        self.header_font = Font(bold=True, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
    def generate_report(self, processed_data, output_path, report_type="standard", include_charts=True):
        """
        Generate Excel report with multiple sheets
        
        Args:
            processed_data (dict): Processed data from DataProcessor
            output_path (str): Path for output Excel file
            report_type (str): Type of report (standard, detailed, summary)
            include_charts (bool): Whether to include charts
        """
        try:
            self.logger.info(f"Generating {report_type} Excel report")
            
            # Create workbook
            wb = openpyxl.Workbook()
            
            # Remove the default worksheet that is created
            wb.remove(wb.worksheets[0])
            
            # Generate sheets based on report type
            if report_type == "summary":
                self._create_summary_sheet(wb, processed_data)
                if include_charts:
                    self._create_charts_sheet(wb, processed_data)
            elif report_type == "detailed":
                self._create_all_sheets(wb, processed_data, include_charts)
            else:  # standard
                self._create_standard_sheets(wb, processed_data, include_charts)
                
            # Save workbook
            wb.save(output_path)
            self.logger.info(f"Report saved successfully to {output_path}")
            
        except Exception as e:
            error_msg = f"Failed to generate Excel report: {str(e)}"
            self.logger.error(error_msg, exc_info=True)
            raise Exception(error_msg)
            
    def _create_summary_sheet(self, wb, data):
        """Create summary overview sheet"""
        ws = wb.create_sheet("Summary")
        
        # Title
        ws['A1'] = "Back Order Report Summary"
        ws['A1'].font = Font(size=16, bold=True)
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        # Summary statistics
        row = 4
        summary = data.get('summary', {})
        
        for key, value in summary.items():
            ws[f'A{row}'] = key.replace('_', ' ').title()
            ws[f'B{row}'] = value
            row += 1
            
        # Format summary section
        self._format_range(ws, f'A4:B{row-1}')
        
        # Top items by quantity
        if 'by_item' in data:
            row += 2
            ws[f'A{row}'] = "Top 10 Items by Quantity"
            ws[f'A{row}'].font = Font(bold=True, size=12)
            row += 1
            
            top_items = data['by_item'].head(10)
            self._write_dataframe(ws, top_items, start_row=row)
            
    def _create_standard_sheets(self, wb, data, include_charts):
        """Create standard report sheets"""
        # Summary sheet
        self._create_summary_sheet(wb, data)
        
        # By Item sheet
        if 'by_item' in data:
            self._create_by_item_sheet(wb, data['by_item'])
            
        # By Customer sheet (if available)
        if 'by_customer' in data:
            self._create_by_customer_sheet(wb, data['by_customer'])
            
        # Aging Analysis
        if 'aging' in data and not data['aging'].empty:
            self._create_aging_sheet(wb, data['aging'])
            
        # Charts
        if include_charts:
            self._create_charts_sheet(wb, data)
            
    def _create_all_sheets(self, wb, data, include_charts):
        """Create all possible sheets for detailed report"""
        # All standard sheets
        self._create_standard_sheets(wb, data, include_charts)
        
        # Additional detailed sheets
        if 'by_supplier' in data:
            self._create_by_supplier_sheet(wb, data['by_supplier'])
            
        if 'by_date' in data and not data['by_date'].empty:
            self._create_by_date_sheet(wb, data['by_date'])
            
        if 'by_category' in data:
            self._create_by_category_sheet(wb, data['by_category'])
            
        # Raw data
        self._create_raw_data_sheet(wb, data['raw_data'])
        
    def _create_by_item_sheet(self, wb, data):
        """Create by item analysis sheet"""
        ws = wb.create_sheet("By Item")
        
        ws['A1'] = "Back Orders by Item"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        self._write_dataframe(ws, data, start_row=4)
        
    def _create_by_customer_sheet(self, wb, data):
        """Create by customer analysis sheet"""
        ws = wb.create_sheet("By Customer")
        
        ws['A1'] = "Back Orders by Customer"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        self._write_dataframe(ws, data, start_row=4)
        
    def _create_by_supplier_sheet(self, wb, data):
        """Create by supplier analysis sheet"""
        ws = wb.create_sheet("By Supplier")
        
        ws['A1'] = "Back Orders by Supplier"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        self._write_dataframe(ws, data, start_row=4)
        
    def _create_by_date_sheet(self, wb, data):
        """Create by date analysis sheet"""
        ws = wb.create_sheet("By Date")
        
        ws['A1'] = "Back Orders by Month"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        self._write_dataframe(ws, data, start_row=4)
        
    def _create_by_category_sheet(self, wb, data):
        """Create by category analysis sheet"""
        ws = wb.create_sheet("By Category")
        
        ws['A1'] = "Back Orders by Category"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        self._write_dataframe(ws, data, start_row=4)
        
    def _create_aging_sheet(self, wb, data):
        """Create aging analysis sheet"""
        ws = wb.create_sheet("Aging Analysis")
        
        ws['A1'] = "Back Order Aging Analysis"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        self._write_dataframe(ws, data, start_row=4)
        
    def _create_raw_data_sheet(self, wb, data):
        """Create raw data sheet"""
        ws = wb.create_sheet("Raw Data")
        
        ws['A1'] = "Raw Back Order Data"
        ws['A1'].font = Font(size=14, bold=True)
        ws['A2'] = f"Generated on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        
        self._write_dataframe(ws, data, start_row=4)
        
    def _create_charts_sheet(self, wb, data):
        """Create charts and visualizations sheet"""
        ws = wb.create_sheet("Charts")
        
        ws['A1'] = "Back Order Analysis Charts"
        ws['A1'].font = Font(size=14, bold=True)
        
        chart_row = 3
        
        # Top items chart
        if 'by_item' in data and not data['by_item'].empty:
            chart_row = self._add_bar_chart(ws, data['by_item'], 'item_code', 'quantity_sum', 
                                          "Top Items by Quantity", chart_row, top_n=10)
            
        # Customer chart
        if 'by_customer' in data and not data['by_customer'].empty:
            chart_row = self._add_bar_chart(ws, data['by_customer'], 'customer', 'quantity_sum',
                                          "Top Customers by Quantity", chart_row, top_n=10)
            
        # Aging pie chart
        if 'aging' in data and not data['aging'].empty:
            self._add_pie_chart(ws, data['aging'], 'age_bucket', 'quantity_sum',
                              "Back Orders by Age", chart_row)
                              
    def _write_dataframe(self, ws, df, start_row=1):
        """Write dataframe to worksheet with formatting"""
        if df.empty:
            ws[f'A{start_row}'] = "No data available"
            return
            
        # Write headers
        for col_idx, column in enumerate(df.columns, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=str(column).replace('_', ' ').title())
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = Alignment(horizontal='center')
            
        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), start_row + 1):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = self.border
                
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
            
    def _format_range(self, ws, range_str):
        """Apply formatting to a range of cells"""
        for row in ws[range_str]:
            for cell in row:
                cell.border = self.border
                
    def _add_bar_chart(self, ws, data, category_col, value_col, title, start_row, top_n=10):
        """Add a bar chart to the worksheet"""
        try:
            # Get top N items
            chart_data = data.head(top_n)
            
            # Write chart data to worksheet
            data_start_row = start_row + 2
            ws.cell(row=start_row, column=1, value=title).font = Font(bold=True)
            
            # Headers
            ws.cell(row=data_start_row, column=1, value=category_col.replace('_', ' ').title())
            ws.cell(row=data_start_row, column=2, value=value_col.replace('_', ' ').title())
            
            # Data
            for idx, row in enumerate(chart_data.itertuples(index=False), data_start_row + 1):
                ws.cell(row=idx, column=1, value=getattr(row, category_col))
                ws.cell(row=idx, column=2, value=getattr(row, value_col))
                
            # Create chart
            chart = BarChart()
            chart.title = title
            chart.x_axis.title = category_col.replace('_', ' ').title()
            chart.y_axis.title = value_col.replace('_', ' ').title()
            
            # Data references
            data_ref = Reference(ws, min_col=2, min_row=data_start_row,
                               max_row=data_start_row + len(chart_data), max_col=2)
            cats_ref = Reference(ws, min_col=1, min_row=data_start_row + 1,
                               max_row=data_start_row + len(chart_data), max_col=1)
                               
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            
            # Position chart
            ws.add_chart(chart, f'D{start_row}')
            
            return start_row + 20
            
        except Exception as e:
            self.logger.warning(f"Failed to create bar chart: {str(e)}")
            return start_row + 5
            
    def _add_pie_chart(self, ws, data, category_col, value_col, title, start_row):
        """Add a pie chart to the worksheet"""
        try:
            # Write chart data to worksheet
            data_start_row = start_row + 2
            ws.cell(row=start_row, column=1, value=title).font = Font(bold=True)
            
            # Headers
            ws.cell(row=data_start_row, column=1, value=category_col.replace('_', ' ').title())
            ws.cell(row=data_start_row, column=2, value=value_col.replace('_', ' ').title())
            
            # Data
            for idx, row in enumerate(data.itertuples(index=False), data_start_row + 1):
                ws.cell(row=idx, column=1, value=getattr(row, category_col))
                ws.cell(row=idx, column=2, value=getattr(row, value_col))
                
            # Create chart
            chart = PieChart()
            chart.title = title
            
            # Data references
            data_ref = Reference(ws, min_col=2, min_row=data_start_row,
                               max_row=data_start_row + len(data), max_col=2)
            cats_ref = Reference(ws, min_col=1, min_row=data_start_row + 1,
                               max_row=data_start_row + len(data), max_col=1)
                               
            chart.add_data(data_ref, titles_from_data=True)
            chart.set_categories(cats_ref)
            
            # Position chart
            ws.add_chart(chart, f'D{start_row}')
            
        except Exception as e:
            self.logger.warning(f"Failed to create pie chart: {str(e)}")
