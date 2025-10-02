#!/usr/bin/env python3

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, numbers, Border, Side
from openpyxl.utils import get_column_letter
from datetime import datetime
import os
import logging
import traceback
import sys
from typing import List, Tuple

class TripleVerbosityErrorLogger:
    """Custom logger for triple verbosity error reporting."""
    
    def __init__(self, log_file: str = None):
        if log_file is None:
            log_file = f"backorder_error_log_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt"
        
        self.log_file = log_file
        self.has_errors = False
        
        # Create custom logger
        self.logger = logging.getLogger('BackorderErrorLogger')
        self.logger.setLevel(logging.DEBUG)
        
        # Remove any existing handlers
        for handler in self.logger.handlers[:]:
            self.logger.removeHandler(handler)
        
        # Console handler for normal operation
        console_handler = logging.StreamHandler()
        console_handler.setLevel(logging.INFO)
        console_format = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
        console_handler.setFormatter(console_format)
        self.logger.addHandler(console_handler)
    
    def log_error(self, error_context: str, exception: Exception, additional_info: dict = None):
        """Log error with triple verbosity to file."""
        self.has_errors = True
        
        # Create file handler only when error occurs
        if not any(isinstance(h, logging.FileHandler) for h in self.logger.handlers):
            file_handler = logging.FileHandler(self.log_file, mode='w', encoding='utf-8')
            file_handler.setLevel(logging.DEBUG)
            file_format = logging.Formatter('%(asctime)s - %(levelname)s - %(message)s')
            file_handler.setFormatter(file_format)
            self.logger.addHandler(file_handler)
        
        # Triple verbosity error logging
        error_msg = f"\n{'='*80}\nERROR REPORT - {error_context}\n{'='*80}\n"
        
        # Level 1: Basic error information
        error_msg += f"LEVEL 1 - BASIC ERROR INFO:\n"
        error_msg += f"  Error Type: {type(exception).__name__}\n"
        error_msg += f"  Error Message: {str(exception)}\n"
        error_msg += f"  Timestamp: {datetime.now().strftime('%Y-%m-%d %H:%M:%S.%f')}\n"
        error_msg += f"  Context: {error_context}\n\n"
        
        # Level 2: Detailed system and environment info
        error_msg += f"LEVEL 2 - SYSTEM & ENVIRONMENT INFO:\n"
        error_msg += f"  Python Version: {sys.version}\n"
        error_msg += f"  Platform: {sys.platform}\n"
        error_msg += f"  Current Working Directory: {os.getcwd()}\n"
        error_msg += f"  Script File: {__file__}\n"
        error_msg += f"  Process ID: {os.getpid()}\n"
        error_msg += f"  User: {os.getenv('USER', 'Unknown')}\n"
        
        # Memory info if available
        try:
            import psutil
            process = psutil.Process()
            error_msg += f"  Memory Usage: {process.memory_info().rss / 1024 / 1024:.2f} MB\n"
        except ImportError:
            error_msg += f"  Memory Usage: Not available (psutil not installed)\n"
        
        error_msg += "\n"
        
        # Level 3: Full stack trace and additional context
        error_msg += f"LEVEL 3 - COMPLETE STACK TRACE & CONTEXT:\n"
        error_msg += f"  Full Stack Trace:\n"
        for line in traceback.format_exception(type(exception), exception, exception.__traceback__):
            error_msg += f"    {line.rstrip()}\n"
        
        # Additional context information
        if additional_info:
            error_msg += f"\n  Additional Context Information:\n"
            for key, value in additional_info.items():
                error_msg += f"    {key}: {value}\n"
        
        # Local variables from the exception frame
        error_msg += f"\n  Local Variables at Error Point:\n"
        tb = exception.__traceback__
        if tb:
            frame = tb.tb_frame
            for var_name, var_value in frame.f_locals.items():
                if not var_name.startswith('__'):
                    try:
                        var_str = str(var_value)[:200]  # Limit length
                        error_msg += f"    {var_name}: {var_str}\n"
                    except:
                        error_msg += f"    {var_name}: <Cannot display value>\n"
        
        error_msg += f"\n{'='*80}\n"
        
        # Log to file
        self.logger.error(error_msg)
        
        # Also log to console
        print(f"❌ ERROR: {error_context} - See {self.log_file} for detailed report")

# Configure logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

class BackorderReportGenerator:
    """Generate backorder reports from Excel data."""
    
    HEADERS = [
        "ORDER #", "CUST PO", "ORDER DATE", "ITEM NO", "MFG", "SHIP ASAP",
        "UNIT PRICE", "UNIT COST", "CUST NAME", "SALESMAN NAME", "DUE DATE",
        "STOCK", "GP UNIT", "GP TOTAL", "TOTAL SALE", "COMMENTS"
    ]
    
    # Column widths in pixels (A through P) - B,D,E,I,N,O increased by 10%
    COLUMN_WIDTHS = [54, 160, 80, 167, 54, 49, 75, 75, 256, 125, 80, 45, 75, 83, 83, 400]
    
    # Legend configuration
    LEGEND_CONFIG = [
        ("GREEN", "SHIPPING TODAY", "00FF00"),
        ("YELLOW", "PROBLEM WITH ORDER -- SEE COMMENTS", "FFFF00"),
        ("RED", "AT TESTING", "FF0000"),
        ("ORANGE", "SCHEDULED ORDER", "FFA500")
    ]
    
    def __init__(self, input_file: str, sort_column: str = "order_no"):
        self.input_file = input_file
        self.sort_column = sort_column
        self.error_logger = TripleVerbosityErrorLogger()
        
    def validate_input_file(self) -> None:
        """Validate that input file exists."""
        try:
            if not os.path.exists(self.input_file):
                raise FileNotFoundError(f"Input file '{self.input_file}' not found.")
            
            # Additional validation
            file_size = os.path.getsize(self.input_file)
            if file_size == 0:
                raise ValueError(f"Input file '{self.input_file}' is empty (0 bytes)")
            
            logger.info(f"Input file validation passed: {self.input_file} ({file_size} bytes)")
            
        except Exception as e:
            additional_info = {
                "input_file_path": self.input_file,
                "current_directory": os.getcwd(),
                "directory_contents": str(os.listdir('.'))[:500]
            }
            self.error_logger.log_error("File Validation", e, additional_info)
            raise
    
    def load_and_clean_data(self) -> pd.DataFrame:
        """Load Excel file and perform initial cleaning."""
        try:
            logger.info(f"Loading data from {self.input_file}")
            
            # Try reading the file
            df = pd.read_excel(self.input_file)
            logger.info(f"Initial data shape: {df.shape}")
            
            # Validate we have enough columns before dropping
            required_columns = max([0, 3, 4, 8, 9, 15, 17, 19]) + 1
            if len(df.columns) < required_columns:
                raise ValueError(f"Input file must have at least {required_columns} columns, but only has {len(df.columns)}")
            
            original_columns = list(df.columns)
            logger.info(f"Original columns: {original_columns}")
            
            # Drop specified columns by index (A,D,E,I,J,P,R,T)
            drop_indices = [0, 3, 4, 8, 9, 15, 17, 19]
            drop_cols = [df.columns[i] for i in drop_indices if i < len(df.columns)]
            df.drop(columns=drop_cols, inplace=True)
            df.reset_index(drop=True, inplace=True)
            
            logger.info(f"Remaining columns after dropping: {list(df.columns)}")
            logger.info(f"Data shape after cleaning: {df.shape}")
            
            # Validate required columns exist
            required_cols = ["due_date", "slsman_nam"]
            missing_cols = [col for col in required_cols if col not in df.columns]
            if missing_cols:
                logger.warning(f"Missing columns: {missing_cols}. Available columns: {list(df.columns)}")
            
            # Convert due_date with error handling
            if "due_date" in df.columns:
                original_due_date_sample = df["due_date"].head().tolist()
                df["due_date"] = pd.to_datetime(df["due_date"], errors="coerce")
                null_dates = df["due_date"].isna().sum()
                if null_dates > 0:
                    logger.warning(f"Found {null_dates} invalid dates that were set to NaT")
                    logger.info(f"Sample original due_date values: {original_due_date_sample}")
            
            return df
            
        except Exception as e:
            additional_info = {
                "input_file": self.input_file,
                "file_size": os.path.getsize(self.input_file) if os.path.exists(self.input_file) else "File not found",
                "pandas_version": pd.__version__,
                "available_excel_engines": ["openpyxl", "xlrd"] if hasattr(pd, 'ExcelFile') else "Unknown"
            }
            
            # Add DataFrame info if it was created
            try:
                if 'df' in locals():
                    additional_info.update({
                        "dataframe_shape": df.shape,
                        "dataframe_columns": list(df.columns),
                        "dataframe_dtypes": str(df.dtypes)
                    })
            except:
                pass
            
            self.error_logger.log_error("Data Loading and Cleaning", e, additional_info)
            raise
    
    
    def split_data_by_salesperson(self, df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """Split data into military (Manuel Ortega + customer filter) and commercial (others)."""
        try:
            if "slsman_nam" not in df.columns or "cust_name" not in df.columns:
                logger.warning("Required columns not found, creating empty DataFrames")
                return pd.DataFrame(), pd.DataFrame()
            
            # Get unique salesperson names for logging
            unique_salespeople = df["slsman_nam"].unique()
            logger.info(f"Unique salesperson names: {unique_salespeople}")
            
            # Normalize to lowercase for case-insensitive matching
            df["cust_name_lower"] = df["cust_name"].str.lower()
            
            # Masks
            manuel_mask = df["slsman_nam"].str.upper() == "MANUEL ORTEGA"
            keyword_mask = df["cust_name_lower"].apply(
                lambda x: any(keyword.lower() in x for keyword in ["DLA", "DFAS", "NAVSUP"]) if isinstance(x, str) else False
            )
            
            # Split
            military_df = df[manuel_mask & keyword_mask].copy()
            commercial_df = df[~(manuel_mask & keyword_mask)].copy()
            
            # Clean up helper column
            for d in [military_df, commercial_df]:
                if "cust_name_lower" in d.columns:
                    d.drop(columns=["cust_name_lower"], inplace=True)
            
            logger.info(f"Military orders: {len(military_df)}, Commercial orders: {len(commercial_df)}")
            return military_df, commercial_df

        except Exception as e:
            additional_info = {
                "dataframe_shape": df.shape,
                "slsman_nam_in_columns": "slsman_nam" in df.columns,
                "cust_name_in_columns": "cust_name" in df.columns,
                "unique_salespeople": df["slsman_nam"].unique().tolist() if "slsman_nam" in df.columns else "Column not found"
            }
            self.error_logger.log_error("Data Splitting by Salesperson", e, additional_info)
            raise

    def sort_dataframes(self, military_df: pd.DataFrame, commercial_df: pd.DataFrame) -> Tuple[pd.DataFrame, pd.DataFrame]:
        """Sort dataframes by the configured sort column."""
        try:
            original_military_shape = military_df.shape
            original_commercial_shape = commercial_df.shape
            
            # Map display names to column names for sorting
            sort_column_map = {
                "order_no": "order_no",
                "slsman_nam": "slsman_nam", 
                "due_date": "due_date",
                "pcx_dock": "pcx_dock"  # This may not exist in current data
            }
            
            actual_sort_column = sort_column_map.get(self.sort_column, "order_no")
            logger.info(f"Sorting dataframes by column: {actual_sort_column}")
            
            if not military_df.empty and actual_sort_column in military_df.columns:
                if actual_sort_column == "order_no":
                    # Check for non-numeric order numbers
                    non_numeric_orders = military_df[pd.to_numeric(military_df[actual_sort_column], errors='coerce').isna()]
                    if not non_numeric_orders.empty:
                        logger.warning(f"Found {len(non_numeric_orders)} non-numeric order numbers in military data")
                elif actual_sort_column == "due_date":
                    # Ensure due_date is datetime for proper sorting
                    military_df[actual_sort_column] = pd.to_datetime(military_df[actual_sort_column], errors="coerce")
                
                military_df = military_df.sort_values(by=actual_sort_column)
                logger.info(f"Military data sorted by {actual_sort_column}: {len(military_df)} rows")
            elif not military_df.empty:
                logger.warning(f"{actual_sort_column} column not found in military data - skipping sort")
            
            if not commercial_df.empty and actual_sort_column in commercial_df.columns:
                if actual_sort_column == "order_no":
                    # Check for non-numeric order numbers
                    non_numeric_orders = commercial_df[pd.to_numeric(commercial_df[actual_sort_column], errors='coerce').isna()]
                    if not non_numeric_orders.empty:
                        logger.warning(f"Found {len(non_numeric_orders)} non-numeric order numbers in commercial data")
                elif actual_sort_column == "due_date":
                    # Ensure due_date is datetime for proper sorting
                    commercial_df[actual_sort_column] = pd.to_datetime(commercial_df[actual_sort_column], errors="coerce")
                
                commercial_df = commercial_df.sort_values(by=actual_sort_column)
                logger.info(f"Commercial data sorted by {actual_sort_column}: {len(commercial_df)} rows")
            elif not commercial_df.empty:
                logger.warning(f"{actual_sort_column} column not found in commercial data - skipping sort")
            
            return military_df, commercial_df

        except Exception as e:
            additional_info = {
                "military_df_shape": original_military_shape,
                "commercial_df_shape": original_commercial_shape,
                "sort_column": self.sort_column,
                "actual_sort_column": actual_sort_column if 'actual_sort_column' in locals() else "N/A",
                "military_has_sort_column": actual_sort_column in military_df.columns if not military_df.empty and 'actual_sort_column' in locals() else "Unknown",
                "commercial_has_sort_column": actual_sort_column in commercial_df.columns if not commercial_df.empty and 'actual_sort_column' in locals() else "Unknown"
            }
            self.error_logger.log_error("DataFrame Sorting", e, additional_info)
            raise


    def deduplicate_commercial_data(self, commercial_df: pd.DataFrame) -> pd.DataFrame:
        """Remove Sara Burrell duplicates if Lisa Miller has identical order (all columns except salesman)."""
        try:
            if commercial_df.empty or "slsman_nam" not in commercial_df.columns:
                return commercial_df
            
            original_shape = commercial_df.shape
            
            lisa_rows = commercial_df[commercial_df["slsman_nam"] == "Lisa Miller"]
            sara_rows = commercial_df[commercial_df["slsman_nam"] == "Sara Burrell"]
            
            if lisa_rows.empty or sara_rows.empty:
                logger.info(f"No duplicates to check - Lisa: {len(lisa_rows)}, Sara: {len(sara_rows)}")
                return commercial_df
            
            # Get all columns except salesman name for comparison
            comparison_cols = [col for col in commercial_df.columns if col != "slsman_nam"]
            logger.info(f"Comparing {len(comparison_cols)} columns for deduplication")
            
            to_remove_idx = []
            for sara_idx, sara_row in sara_rows.iterrows():
                # Check if any Lisa row matches all columns except salesman name
                for lisa_idx, lisa_row in lisa_rows.iterrows():
                    # Compare all non-salesman columns
                    match = True
                    for col in comparison_cols:
                        if sara_row[col] != lisa_row[col]:
                            # Handle NaN comparisons
                            if pd.isna(sara_row[col]) and pd.isna(lisa_row[col]):
                                continue
                            match = False
                            break
                    
                    if match:
                        logger.info(f"Found duplicate: Sara row {sara_idx} matches Lisa row {lisa_idx}")
                        logger.info(f"  Order: {sara_row.get('order_no', 'N/A')}, Item: {sara_row.get('item_no', 'N/A')}")
                        to_remove_idx.append(sara_idx)
                        break  # Found a match, no need to check other Lisa rows
            
            if to_remove_idx:
                commercial_df = commercial_df.drop(index=to_remove_idx)
                logger.info(f"Removed {len(to_remove_idx)} duplicate Sara Burrell entries")
            else:
                logger.info("No Sara Burrell duplicates found")
            
            logger.info(f"Deduplication complete: {original_shape} -> {commercial_df.shape}")
            return commercial_df
            
        except Exception as e:
            additional_info = {
                "commercial_df_shape": commercial_df.shape if not commercial_df.empty else "Empty DataFrame",
                "has_slsman_nam_column": "slsman_nam" in commercial_df.columns if not commercial_df.empty else "Empty DataFrame",
                "unique_salespeople": commercial_df["slsman_nam"].unique().tolist() if not commercial_df.empty and "slsman_nam" in commercial_df.columns else "N/A",
                "lisa_count": len(commercial_df[commercial_df["slsman_nam"] == "Lisa Miller"]) if not commercial_df.empty and "slsman_nam" in commercial_df.columns else 0,
                "sara_count": len(commercial_df[commercial_df["slsman_nam"] == "Sara Burrell"]) if not commercial_df.empty and "slsman_nam" in commercial_df.columns else 0
            }
            self.error_logger.log_error("Commercial Data Deduplication", e, additional_info)
            raise
    
    def add_data_to_sheet(self, ws, data: pd.DataFrame) -> None:
        """Add data rows to worksheet."""
        try:
            expected_cols = [
                "order_no", "cust_po", "order_dt", "item_no", "manu_no", "ship_asap",
                "unit_price", "unit_cost", "cust_name", "slsman_nam", "due_date", "from_stk"
            ]
            
            logger.info(f"Adding {len(data)} rows to worksheet")
            
            for row_idx, (_, row) in enumerate(data.iterrows()):
                row_data = []
                for col in expected_cols:
                    if col in data.columns:
                        row_data.append(row[col])
                    else:
                        row_data.append("")  # Default for missing columns
                
                # Add calculated columns (will be filled with formulas later)
                row_data.extend([None, None, None, ""])  # GP UNIT, GP TOTAL, TOTAL SALE, COMMENTS
                ws.append(row_data)
            
            logger.info(f"Successfully added {len(data)} data rows to worksheet")
            
        except Exception as e:
            additional_info = {
                "data_shape": data.shape,
                "data_columns": list(data.columns),
                "expected_columns": expected_cols,
                "worksheet_name": ws.title if hasattr(ws, 'title') else "Unknown",
                "current_row_being_processed": row_idx if 'row_idx' in locals() else "N/A"
            }
            self.error_logger.log_error("Adding Data to Worksheet", e, additional_info)
            raise
    
    def format_sheet_headers(self, ws) -> None:
        """Apply formatting to sheet headers."""
        try:
            for col in range(1, len(self.HEADERS) + 1):
                cell = ws[f"{get_column_letter(col)}1"]
                cell.font = Font(bold=True, underline="single")
                cell.alignment = Alignment(horizontal="center")
            
            logger.info(f"Applied header formatting to {len(self.HEADERS)} columns")
            
        except Exception as e:
            additional_info = {
                "headers_count": len(self.HEADERS),
                "headers": self.HEADERS,
                "worksheet_name": ws.title if hasattr(ws, 'title') else "Unknown"
            }
            self.error_logger.log_error("Header Formatting", e, additional_info)
            raise
    
    def set_column_widths(self, ws) -> None:
        """Set column widths in pixels."""
        try:
            for col, width_pixels in enumerate(self.COLUMN_WIDTHS, start=1):
                # Convert pixels to Excel column width units (approximately 7 pixels per unit)
                width_units = width_pixels / 7
                ws.column_dimensions[get_column_letter(col)].width = width_units
            
            logger.info(f"Set column widths for {len(self.COLUMN_WIDTHS)} columns")
            
        except Exception as e:
            additional_info = {
                "column_widths": self.COLUMN_WIDTHS,
                "worksheet_name": ws.title if hasattr(ws, 'title') else "Unknown"
            }
            self.error_logger.log_error("Column Width Setting", e, additional_info)
            raise
    
    def add_formulas_and_formatting(self, ws) -> None:
        """Add Excel formulas and number formatting."""
        try:
            max_row = ws.max_row
            logger.info(f"Adding formulas and formatting to {max_row - 1} data rows")
            
            for row in range(2, max_row + 1):
                # Add formulas
                ws[f"M{row}"] = f"=G{row}-H{row}"  # GP UNIT
                ws[f"N{row}"] = f"=M{row}*L{row}"  # GP TOTAL
                ws[f"O{row}"] = f"=L{row}*G{row}"  # TOTAL SALE
                
                # Apply alignment
                for col in range(1, len(self.HEADERS) + 1):
                    ws[f"{get_column_letter(col)}{row}"].alignment = Alignment(horizontal="right")
                
                # Apply currency formatting
                for col in ["G", "H", "M", "N", "O"]:
                    cell = ws[f"{col}{row}"]
                    cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
                
                # Apply date formatting to columns C and K (MM-DD-YY format)
                for col in ["C", "K"]:
                    cell = ws[f"{col}{row}"]
                    cell.number_format = "MM-DD-YY"
            
            logger.info("Successfully applied formulas and formatting")
            
        except Exception as e:
            additional_info = {
                "worksheet_max_row": ws.max_row,
                "worksheet_name": ws.title if hasattr(ws, 'title') else "Unknown",
                "current_row_being_processed": row if 'row' in locals() else "N/A"
            }
            self.error_logger.log_error("Formula and Formatting Application", e, additional_info)
            raise
    
    def add_totals_row(self, ws) -> int:
        """Add totals row and return the row number."""
        try:
            max_row = ws.max_row
            total_row = max_row + 3
            
            # Add total formulas
            ws[f"N{total_row}"] = f"=SUM(N2:N{max_row})"
            ws[f"O{total_row}"] = f"=SUM(O2:O{max_row})"
            
            # Format totals
            for col in ["N", "O"]:
                cell = ws[f"{col}{total_row}"]
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="right")
                cell.number_format = numbers.FORMAT_CURRENCY_USD_SIMPLE
            
            logger.info(f"Added totals row at row {total_row}")
            return total_row
            
        except Exception as e:
            additional_info = {
                "worksheet_max_row": ws.max_row,
                "calculated_total_row": max_row + 3 if 'max_row' in locals() else "N/A",
                "worksheet_name": ws.title if hasattr(ws, 'title') else "Unknown"
            }
            self.error_logger.log_error("Totals Row Addition", e, additional_info)
            raise
    
    def add_legend(self, ws, start_row: int) -> None:
        """Add color legend to worksheet."""
        try:
            legend_start = start_row + 3
            thick_border = Border(
                left=Side(border_style="thick", color="000000"),
                right=Side(border_style="thick", color="000000"),
                top=Side(border_style="thick", color="000000"),
                bottom=Side(border_style="thick", color="000000")
            )
            
            # Legend headers
            ws[f"A{legend_start}"] = "COLOR"
            ws[f"B{legend_start}"] = "MEANING"
            
            # Merge B header with C,D,E,F and center
            ws.merge_cells(f"B{legend_start}:F{legend_start}")
            
            for cell in [ws[f"A{legend_start}"], ws[f"B{legend_start}"]]:
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
            
            # Legend entries
            for i, (color, meaning, hex_color) in enumerate(self.LEGEND_CONFIG, start=1):
                row_num = legend_start + i
                color_cell = ws[f"A{row_num}"]
                meaning_cell = ws[f"B{row_num}"]
                
                # Set color cell properties
                color_cell.value = color
                color_cell.fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")
                color_cell.alignment = Alignment(horizontal="center")
                color_cell.border = thick_border
                
                # Merge meaning cell with columns C,D,E,F
                ws.merge_cells(f"B{row_num}:F{row_num}")
                meaning_cell.value = meaning
                meaning_cell.alignment = Alignment(horizontal="center")
                
                # Apply thick border to all cells in the merged range B,C,D,E,F
                for col in ['B', 'C', 'D', 'E', 'F']:
                    ws[f"{col}{row_num}"].border = thick_border
            
            logger.info(f"Added legend starting at row {legend_start}")
            
        except Exception as e:
            additional_info = {
                "legend_start_row": legend_start if 'legend_start' in locals() else start_row + 3,
                "legend_config": self.LEGEND_CONFIG,
                "worksheet_name": ws.title if hasattr(ws, 'title') else "Unknown"
            }
            self.error_logger.log_error("Legend Addition", e, additional_info)
            raise
    
    def create_sheet(self, wb: Workbook, title: str, data: pd.DataFrame) -> None:
        """Create a worksheet with data and formatting."""
        try:
            logger.info(f"Creating sheet '{title}' with {len(data)} rows")
            
            ws = wb.create_sheet(title)
            ws.append(self.HEADERS)
            
            if not data.empty:
                self.add_data_to_sheet(ws, data)
            
            self.format_sheet_headers(ws)
            self.set_column_widths(ws)
            ws.freeze_panes = "A2"
            
            if not data.empty:
                self.add_formulas_and_formatting(ws)
                total_row = self.add_totals_row(ws)
                self.add_legend(ws, total_row)
            
            logger.info(f"Successfully created sheet '{title}'")
            
        except Exception as e:
            additional_info = {
                "sheet_title": title,
                "data_shape": data.shape,
                "data_empty": data.empty,
                "workbook_sheet_count": len(wb.sheetnames)
            }
            self.error_logger.log_error(f"Sheet Creation - {title}", e, additional_info)
            raise
    
    def generate_report(self, output_file_path: str = None, sort_column: str = None) -> str:
        """Generate the complete backorder report."""
        try:
            # Update sort column if provided
            if sort_column:
                self.sort_column = sort_column
                
            logger.info(f"Starting backorder report generation with sort column: {self.sort_column}")
            
            self.validate_input_file()
            
            # Load and process data
            df = self.load_and_clean_data()
            military_df, commercial_df = self.split_data_by_salesperson(df)
            military_df, commercial_df = self.sort_dataframes(military_df, commercial_df)
            commercial_df = self.deduplicate_commercial_data(commercial_df)
            
            # Create workbook
            wb = Workbook()
            wb.remove(wb.active)  # Remove default sheet
            
            self.create_sheet(wb, "MILITARY", military_df)
            self.create_sheet(wb, "COMMERCIAL", commercial_df)
            
            # Save file
            if output_file_path:
                output_file = output_file_path
            else:
                output_file = f"BACKORDER REPORT {datetime.today().strftime('%m%d%y')}.xlsx"
            wb.save(output_file)
            
            logger.info(f"Report generated successfully: {output_file}")
            
            # Log summary
            logger.info(f"Report Summary:")
            logger.info(f"  Total rows processed: {len(df)}")
            logger.info(f"  Military orders: {len(military_df)}")
            logger.info(f"  Commercial orders: {len(commercial_df)}")
            logger.info(f"  Output file: {output_file}")
            logger.info(f"  File size: {os.path.getsize(output_file)} bytes")
            
            return output_file
            
        except Exception as e:
            additional_info = {
                "input_file": self.input_file,
                "current_working_directory": os.getcwd(),
                "directory_listing": str(os.listdir('.'))[:500],
                "processing_stage": "Report Generation Main Process"
            }
            
            # Try to add more context if variables exist
            try:
                if 'df' in locals():
                    additional_info["original_data_shape"] = df.shape
                if 'military_df' in locals():
                    additional_info["military_data_shape"] = military_df.shape
                if 'commercial_df' in locals():
                    additional_info["commercial_data_shape"] = commercial_df.shape
                if 'wb' in locals():
                    additional_info["workbook_sheets"] = wb.sheetnames
            except:
                pass
            
            self.error_logger.log_error("Report Generation", e, additional_info)
            raise

def main():
    """Main function to run the report generator."""
    input_file = "back orders by salesperson report.xls"
    error_logger = TripleVerbosityErrorLogger()
    
    try:
        logger.info("="*60)
        logger.info("BACKORDER REPORT GENERATOR STARTING")
        logger.info("="*60)
        logger.info(f"Input file: {input_file}")
        logger.info(f"Start time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
        
        generator = BackorderReportGenerator(input_file)
        output_file = generator.generate_report()
        
        logger.info("="*60)
        logger.info("BACKORDER REPORT GENERATOR COMPLETED SUCCESSFULLY")
        logger.info("="*60)
        print(f"✅ Report generated successfully: {output_file}")
        
        # Only mention error log if errors occurred
        if generator.error_logger.has_errors:
            print(f"⚠️  Errors occurred during processing - see {generator.error_logger.log_file} for details")
        else:
            logger.info("No errors occurred during processing - no error log created")
        
    except Exception as e:
        additional_info = {
            "input_file": input_file,
            "main_function_stage": "Main execution",
            "current_time": datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            "python_executable": sys.executable,
            "command_line_args": sys.argv
        }
        
        error_logger.log_error("Main Function Execution", e, additional_info)
        
        logger.error("="*60)
        logger.error("BACKORDER REPORT GENERATOR FAILED")
        logger.error("="*60)
        print(f"❌ Error: {e}")
        print(f"📋 Detailed error report saved to: {error_logger.log_file}")
        return 1
    
    return 0

if __name__ == "__main__":
    import sys
    sys.exit(main())